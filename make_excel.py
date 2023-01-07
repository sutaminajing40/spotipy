import datetime
import openpyxl 
import os
import shutil

path = '/Users/iguchihiroto/Documents/programming/spotipy/evaluation.xlsx'
if os.path.isfile(path):
    wb = openpyxl.load_workbook(path)
else:
    wb = openpyxl.Workbook()
ws = wb.active

#初期設定 columnsに列名 indexに行名 どちらも省略可能
def initial_setting(columns:list=[],index:list=[]):
    #1行に列名を書き込み
    for i in range(0,len(columns)):
        ws.cell(1,i+1,value = columns[i])

    #columnsが設定されてなかったら1行目から
    #columnsが設定されているときは2行目から
    if len(columns) == 0:
        cnt = 1
    else:
        cnt = 2

    for i in range(0,len(index)):
        ws.cell(i+cnt,1,value = index[i])
    

def make_backup(file_name:str):
    shutil.copyfile(file_name,file_name+'_backup/'+str(datetime.datetime.now().strftime('%Y年%m月%d日 %H時%M分%S秒'))+file_name)

#wr_mode:index 行で書き込み columns 列で書き込み
def write(list:list,wr_mode='index'):

    max_row = ws.max_row+1
    max_column = ws.max_column+1

    if wr_mode == 'index':
        for i in range(0,len(list)):
            #A列に列名を書き込み
            ws.cell(max_row,i+1,value = list[i])


    if wr_mode == 'columns':
        for i in range(0,len(list)):
            #A列に列名を書き込み
            ws.cell(i+1,max_column,value = list[i])
    
    wb.save('evaluation.xlsx')
    



if __name__ == '__main__':
    #列名
    columns = ['推薦元','アーティスト名','->','推薦結果','アーティスト名','評価']
    initial_setting(columns = columns)
    #write(['サウダージ','ポルノグラフィティ','','ジョバイロ','ポルノグラフィティ',''])

    wb.save('evaluation.xlsx')